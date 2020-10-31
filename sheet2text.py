#! python
# sheet2textpy - script read every column from the current worksheet and copy it to different .txt files
# X 2020 Arnold Cytrowski



import openpyxl
from openpyxl.utils import get_column_letter

choice = input('give me the name of .xlsx file from current directory tree:\n')

if not choice.endswith('.xlsx'):
    print('it isn\'t the right file, sorry')
    exit()





wb = openpyxl.load_workbook(choice)


sheet = wb.active

max_col = sheet.max_column
max_row = sheet.max_row

for col_num in range(1, max_col + 1):
    text = []
    letter = get_column_letter(col_num)
    file = open(f'{letter}.txt', 'w')
    for row_num in range(1, max_row + 1):
        value = sheet.cell(row = row_num, column = col_num).value
        if value is not None:
            file.write(sheet.cell(row = row_num, column = col_num).value + '\n')
    file.close()

print('aaand it\'s done, thank youuuuu')
        
        







